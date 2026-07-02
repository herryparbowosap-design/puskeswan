"""Modul Stok — obat & alat kesehatan. Berbasis KARTU STOK (transaksi):
saldo = akumulasi transaksi, bukan angka yang diedit langsung.

Jenis transaksi:
  - masuk        : penerimaan (pembelian/droping/hibah). Boleh batch + kedaluwarsa.
  - keluar       : pemakaian/rusak/kedaluwarsa/dibuang (manual). (Pengurangan
                   otomatis dari pelayanan menyusul di slice berikutnya.)
  - penyesuaian  : koreksi (mis. hasil opname). jumlah bertanda (bisa +/-).

Setiap transaksi menyimpan `mutasi` (efek bertanda ke saldo) agar saldo = Σ mutasi.
Slice ini melacak SALDO TOTAL per item; batch/ED dicatat sebagai metadata + daftar
"mendekati kedaluwarsa" (pelacakan sisa per-batch/FEFO menyusul).
"""
import uuid
from collections import defaultdict
from datetime import datetime, timezone, date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel

from db import get_db
from auth import require_roles

router = APIRouter(prefix="/stok", tags=["stok"])

TIPE = {"obat", "alkes"}
JENIS = {"masuk", "keluar", "penyesuaian"}


class ItemIn(BaseModel):
    tipe: str = "obat"
    nama: str
    satuan: str
    obat_id: Optional[str] = None          # tautan opsional ke formularium obat
    stok_minimum: Optional[float] = None    # ambang peringatan stok menipis
    produsen: Optional[str] = None
    kemasan: Optional[str] = None           # Botol / box / dll
    sediaan: Optional[str] = None           # 100 ml / 5 ml / ukuran
    kategori: Optional[str] = None          # obat: Multivitamin/Antibiotik/...
    gauge: Optional[str] = None             # alkes: ukuran jarum (18G/21G)
    keterangan: Optional[str] = None        # mis. "1 box @ 100 pcs"
    aktif: bool = True


def _item_doc_fields(body):
    return {
        "produsen": body.produsen, "kemasan": body.kemasan, "sediaan": body.sediaan,
        "kategori": body.kategori, "gauge": body.gauge, "keterangan": body.keterangan,
    }


class TransaksiIn(BaseModel):
    item_id: str
    jenis: str                              # masuk | keluar | penyesuaian
    jumlah: float                           # masuk/keluar: magnitudo; penyesuaian: bertanda
    tgl: Optional[str] = None
    batch: Optional[str] = None
    exp: Optional[str] = None               # ISO YYYY-MM-DD (obat masuk)
    sumber: Optional[str] = None            # pembelian/droping/hibah/pemakaian/rusak/kedaluwarsa/opname
    catatan: Optional[str] = None


async def _saldo_map(db, item_ids=None):
    flt = {"item_id": {"$in": item_ids}} if item_ids is not None else {}
    txs = await db.stok_transaksi.find(flt, {"_id": 0, "item_id": 1, "mutasi": 1}).to_list(100000)
    m = defaultdict(float)
    for t in txs:
        m[t["item_id"]] += t.get("mutasi", 0) or 0
    return m


def _rapikan(x):
    """Tampilkan bilangan bulat tanpa .0."""
    try:
        f = float(x)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return x


async def potong_stok_pelayanan(db, pelayanan_id, tgl, obat_list, user):
    """Potong stok dari obat yang dipakai di pelayanan (best-effort, non-blocking).
    Hanya memotong bila item stok obat cocok (via obat_id atau nama) DAN satuan sama.
    Selain itu dicatat sebagai 'dilewati' agar transparan (mis. satuan beda ml vs Botol)."""
    dipotong, dilewati = [], []
    now = datetime.now(timezone.utc)
    for o in obat_list or []:
        try:
            jml = float(o.get("jumlah") or 0)
        except (TypeError, ValueError):
            jml = 0
        if jml <= 0:
            continue
        nama = o.get("nama")
        item = None
        if o.get("obat_id"):
            item = await db.stok_item.find_one({"tipe": "obat", "obat_id": o["obat_id"]}, {"_id": 0})
        if not item and nama:
            item = await db.stok_item.find_one({"tipe": "obat", "nama": nama}, {"_id": 0})
        if not item:
            dilewati.append({"nama": nama, "alasan": "item stok tidak ada"})
            continue
        su_item = (item.get("satuan") or "").strip().lower()
        su_pakai = (o.get("satuan") or "").strip().lower()
        if su_pakai and su_item and su_pakai != su_item:
            dilewati.append({"nama": nama, "alasan": f"satuan beda ({o.get('satuan')} vs {item.get('satuan')})"})
            continue
        await db.stok_transaksi.insert_one({
            "id": uuid.uuid4().hex, "item_id": item["id"], "item_nama": item["nama"], "satuan": item.get("satuan"),
            "jenis": "keluar", "mutasi": -jml, "tgl": tgl or now.date().isoformat(),
            "batch": None, "exp": None, "sumber": "pelayanan", "catatan": "pemakaian pelayanan",
            "ref_pelayanan": pelayanan_id, "petugas_id": user.get("id"), "created_at": now,
        })
        dipotong.append({"nama": item["nama"], "jumlah": _rapikan(jml), "satuan": item.get("satuan")})
    return {"dipotong": dipotong, "dilewati": dilewati}


async def kembalikan_stok_pelayanan(db, pelayanan_id):
    """Hapus transaksi keluar yang berasal dari pelayanan (dipakai saat pelayanan dihapus)."""
    r = await db.stok_transaksi.delete_many({"ref_pelayanan": pelayanan_id})
    return r.deleted_count


# --------------------------------------------------------------- ITEM
@router.get("/item")
async def list_item(
    tipe: Optional[str] = None,
    _user=Depends(require_roles("petugas", "admin")),
):
    db = get_db()
    flt = {}
    if tipe in TIPE:
        flt["tipe"] = tipe
    items = await db.stok_item.find(flt, {"_id": 0}).sort("nama", 1).to_list(2000)
    saldo = await _saldo_map(db)
    for it in items:
        s = saldo.get(it["id"], 0)
        it["saldo"] = _rapikan(s)
        it["stok_rendah"] = it.get("stok_minimum") is not None and s <= it["stok_minimum"]
    return items


@router.post("/item")
async def create_item(body: ItemIn, user=Depends(require_roles("petugas", "admin"))):
    if body.tipe not in TIPE:
        raise HTTPException(400, "tipe harus obat/alkes")
    if not body.nama.strip() or not body.satuan.strip():
        raise HTTPException(400, "nama & satuan wajib")
    db = get_db()
    now = datetime.now(timezone.utc)
    doc = {
        "id": uuid.uuid4().hex,
        "tipe": body.tipe,
        "nama": body.nama.strip(),
        "satuan": body.satuan.strip(),
        "obat_id": body.obat_id,
        "stok_minimum": body.stok_minimum,
        **_item_doc_fields(body),
        "aktif": body.aktif,
        "created_by": user["id"],
        "created_at": now,
    }
    await db.stok_item.insert_one(doc)
    doc.pop("_id", None)
    doc["saldo"] = 0
    doc["stok_rendah"] = False
    return doc


@router.patch("/item/{iid}")
async def edit_item(iid: str, body: ItemIn, _user=Depends(require_roles("petugas", "admin"))):
    if body.tipe not in TIPE:
        raise HTTPException(400, "tipe harus obat/alkes")
    r = await get_db().stok_item.update_one({"id": iid}, {"$set": {
        "tipe": body.tipe, "nama": body.nama.strip(), "satuan": body.satuan.strip(),
        "obat_id": body.obat_id, "stok_minimum": body.stok_minimum,
        **_item_doc_fields(body),
        "aktif": body.aktif,
    }})
    if r.matched_count == 0:
        raise HTTPException(404, "item tidak ditemukan")
    return {"ok": True}


@router.delete("/item/{iid}")
async def delete_item(iid: str, cascade: bool = False, _user=Depends(require_roles("admin"))):
    db = get_db()
    n_tx = await db.stok_transaksi.count_documents({"item_id": iid})
    if n_tx and not cascade:
        raise HTTPException(409, f"item punya {n_tx} transaksi; hapus dengan ?cascade=true untuk force")
    if cascade:
        await db.stok_transaksi.delete_many({"item_id": iid})
    r = await db.stok_item.delete_one({"id": iid})
    if r.deleted_count == 0:
        raise HTTPException(404, "item tidak ditemukan")
    return {"ok": True, "transaksi_terhapus": n_tx if cascade else 0}


@router.get("/item/{iid}")
async def detail_item(iid: str, _user=Depends(require_roles("petugas", "admin"))):
    db = get_db()
    it = await db.stok_item.find_one({"id": iid}, {"_id": 0})
    if not it:
        raise HTTPException(404, "item tidak ditemukan")
    saldo = (await _saldo_map(db, [iid])).get(iid, 0)
    it["saldo"] = _rapikan(saldo)
    it["stok_rendah"] = it.get("stok_minimum") is not None and saldo <= it["stok_minimum"]
    txs = await db.stok_transaksi.find({"item_id": iid}, {"_id": 0}).sort("tgl", -1).limit(200).to_list(200)
    return {"item": it, "transaksi": txs}


# --------------------------------------------------------------- TRANSAKSI
@router.post("/transaksi")
async def create_transaksi(body: TransaksiIn, user=Depends(require_roles("petugas", "admin"))):
    if body.jenis not in JENIS:
        raise HTTPException(400, "jenis harus masuk/keluar/penyesuaian")
    db = get_db()
    it = await db.stok_item.find_one({"id": body.item_id}, {"_id": 0, "id": 1, "nama": 1, "satuan": 1})
    if not it:
        raise HTTPException(400, "item tidak ditemukan")
    if body.jenis == "masuk":
        if body.jumlah <= 0:
            raise HTTPException(400, "jumlah masuk harus > 0")
        mutasi = abs(body.jumlah)
    elif body.jenis == "keluar":
        if body.jumlah <= 0:
            raise HTTPException(400, "jumlah keluar harus > 0")
        mutasi = -abs(body.jumlah)
    else:  # penyesuaian — bertanda
        mutasi = body.jumlah
    now = datetime.now(timezone.utc)
    doc = {
        "id": uuid.uuid4().hex,
        "item_id": body.item_id,
        "item_nama": it["nama"],
        "satuan": it.get("satuan"),
        "jenis": body.jenis,
        "mutasi": mutasi,
        "tgl": body.tgl or now.date().isoformat(),
        "batch": body.batch,
        "exp": body.exp,
        "sumber": body.sumber,
        "catatan": body.catatan,
        "petugas_id": user["id"],
        "created_at": now,
    }
    await db.stok_transaksi.insert_one(doc)
    doc.pop("_id", None)
    saldo = (await _saldo_map(db, [body.item_id])).get(body.item_id, 0)
    return {"transaksi": doc, "saldo": _rapikan(saldo), "saldo_minus": saldo < 0}


@router.delete("/transaksi/{tid}")
async def delete_transaksi(tid: str, _user=Depends(require_roles("admin"))):
    r = await get_db().stok_transaksi.delete_one({"id": tid})
    if r.deleted_count == 0:
        raise HTTPException(404, "transaksi tidak ditemukan")
    return {"ok": True}


# --------------------------------------------------------------- KEDALUWARSA
async def _mendekati_ed(db, hari=90):
    batas = (date.today() + timedelta(days=hari)).isoformat()
    hari_ini = date.today().isoformat()
    rows = await db.stok_transaksi.find(
        {"jenis": "masuk", "exp": {"$ne": None, "$lte": batas}}, {"_id": 0}
    ).sort("exp", 1).to_list(500)
    for r in rows:
        r["kedaluwarsa"] = bool(r.get("exp") and r["exp"] < hari_ini)
    return rows


@router.get("/kedaluwarsa")
async def mendekati_kedaluwarsa(
    hari: int = Query(90, ge=1, le=730),
    _user=Depends(require_roles("petugas", "admin")),
):
    """Penerimaan (masuk) dengan tanggal ED dalam `hari` ke depan (info; belum FEFO)."""
    return await _mendekati_ed(get_db(), hari)


# --------------------------------------------------------------- LAPORAN
def _rekap_periode_map(txs):
    masuk, keluar, sesuai = defaultdict(float), defaultdict(float), defaultdict(float)
    for t in txs:
        j = t.get("jenis")
        m = t.get("mutasi", 0) or 0
        if j == "masuk":
            masuk[t["item_id"]] += m
        elif j == "keluar":
            keluar[t["item_id"]] += -m
        elif j == "penyesuaian":
            sesuai[t["item_id"]] += m
    return masuk, keluar, sesuai


async def _laporan_data(db, tahun=None, bulan=None, hari_ed=90):
    items = await db.stok_item.find({}, {"_id": 0}).sort("nama", 1).to_list(2000)
    saldo = await _saldo_map(db)
    rekap_masuk = rekap_keluar = rekap_sesuai = {}
    periode = None
    if tahun and bulan:
        s_dari = f"{tahun:04d}-{bulan:02d}-01"
        s_sampai = f"{tahun:04d}-{bulan + 1:02d}-01" if bulan < 12 else f"{tahun + 1:04d}-01-01"
        txs = await db.stok_transaksi.find({"tgl": {"$gte": s_dari, "$lt": s_sampai}}, {"_id": 0}).to_list(100000)
        rekap_masuk, rekap_keluar, rekap_sesuai = _rekap_periode_map(txs)
        periode = {"tahun": tahun, "bulan": bulan}
    daftar = []
    n_menipis = 0
    for it in items:
        s = saldo.get(it["id"], 0)
        rendah = it.get("stok_minimum") is not None and s <= it["stok_minimum"]
        if rendah:
            n_menipis += 1
        daftar.append({
            "id": it["id"], "tipe": it["tipe"], "nama": it["nama"], "satuan": it.get("satuan"),
            "produsen": it.get("produsen"), "kategori": it.get("kategori"), "sediaan": it.get("sediaan"),
            "saldo": _rapikan(s), "stok_minimum": it.get("stok_minimum"), "stok_rendah": rendah,
            "masuk": _rapikan(rekap_masuk.get(it["id"], 0)) if periode else None,
            "keluar": _rapikan(rekap_keluar.get(it["id"], 0)) if periode else None,
            "penyesuaian": _rapikan(rekap_sesuai.get(it["id"], 0)) if periode else None,
        })
    ed = await _mendekati_ed(db, hari_ed)
    return {
        "periode": periode,
        "ringkas": {"item": len(items), "menipis": n_menipis, "mendekati_ed": len(ed)},
        "saldo": daftar,
        "mendekati_ed": ed,
    }


@router.get("/laporan")
async def laporan_stok(
    tahun: Optional[int] = None, bulan: Optional[int] = None, hari_ed: int = 90,
    _user=Depends(require_roles("petugas", "admin")),
):
    return await _laporan_data(get_db(), tahun, bulan, hari_ed)


@router.get("/laporan/xlsx")
async def laporan_stok_xlsx(
    tahun: Optional[int] = None, bulan: Optional[int] = None, hari_ed: int = 90,
    _user=Depends(require_roles("petugas", "admin")),
):
    from fastapi.responses import Response
    data = await _laporan_data(get_db(), tahun, bulan, hari_ed)
    blob = _build_stok_xlsx(data)
    fname = "laporan-stok"
    if data["periode"]:
        fname += f"-{data['periode']['tahun']:04d}-{data['periode']['bulan']:02d}"
    return Response(
        content=blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
    )


def _build_stok_xlsx(data):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="0F6E56")
    TITLE = Font(bold=True, size=14)

    def hdr(ws, row, cols):
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=j, value=c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL

    def autow(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 45)

    periode = data.get("periode")
    wb = Workbook()
    ws = wb.active
    ws.title = "Saldo"
    ws["A1"] = "LAPORAN STOK" + (f" — {periode['bulan']:02d}/{periode['tahun']}" if periode else "")
    ws["A1"].font = TITLE
    kolom = ["Tipe", "Nama", "Produsen", "Kategori", "Sediaan", "Satuan", "Saldo", "Min", "Status"]
    if periode:
        kolom += ["Masuk", "Keluar", "Penyesuaian"]
    hdr(ws, 3, kolom)
    r = 4
    for it in data["saldo"]:
        row = [it["tipe"], it["nama"], it.get("produsen"), it.get("kategori"), it.get("sediaan"),
               it.get("satuan"), it["saldo"], it.get("stok_minimum"), "menipis" if it["stok_rendah"] else ""]
        if periode:
            row += [it.get("masuk"), it.get("keluar"), it.get("penyesuaian")]
        for j, v in enumerate(row, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    autow(ws)

    ws2 = wb.create_sheet("Mendekati ED")
    hdr(ws2, 1, ["Item", "Satuan", "Jumlah masuk", "Tgl ED", "Status"])
    for i, e in enumerate(data["mendekati_ed"], start=2):
        ws2.cell(row=i, column=1, value=e.get("item_nama"))
        ws2.cell(row=i, column=2, value=e.get("satuan"))
        ws2.cell(row=i, column=3, value=_rapikan(e.get("mutasi", 0)))
        ws2.cell(row=i, column=4, value=e.get("exp"))
        ws2.cell(row=i, column=5, value="LEWAT ED" if e.get("kedaluwarsa") else "mendekati")
    autow(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------- OPNAME
class OpnameItemIn(BaseModel):
    item_id: str
    fisik: Optional[float] = None


class OpnamePatchIn(BaseModel):
    items: list[OpnameItemIn] = []
    catatan: Optional[str] = None


@router.post("/opname")
async def mulai_opname(tipe: Optional[str] = None, user=Depends(require_roles("petugas", "admin"))):
    """Mulai sesi opname: snapshot stok sistem untuk item aktif (status draft)."""
    db = get_db()
    flt = {"aktif": {"$ne": False}}
    if tipe in TIPE:
        flt["tipe"] = tipe
    items = await db.stok_item.find(flt, {"_id": 0, "id": 1, "nama": 1, "satuan": 1, "tipe": 1}).sort("nama", 1).to_list(2000)
    saldo = await _saldo_map(db, [i["id"] for i in items])
    now = datetime.now(timezone.utc)
    doc = {
        "id": uuid.uuid4().hex,
        "tgl": now.date().isoformat(),
        "status": "draft",
        "tipe": tipe if tipe in TIPE else "semua",
        "catatan": None,
        "items": [{
            "item_id": i["id"], "nama": i["nama"], "satuan": i.get("satuan"), "tipe": i.get("tipe"),
            "sistem_awal": _rapikan(saldo.get(i["id"], 0)), "fisik": None,
        } for i in items],
        "petugas_id": user["id"],
        "created_at": now,
        "selesai_at": None,
    }
    await db.stok_opname.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.get("/opname")
async def list_opname(_user=Depends(require_roles("petugas", "admin"))):
    rows = await get_db().stok_opname.find({}, {"_id": 0, "items": 0}).sort("created_at", -1).limit(100).to_list(100)
    return rows


@router.get("/opname/{oid}")
async def detail_opname(oid: str, _user=Depends(require_roles("petugas", "admin"))):
    o = await get_db().stok_opname.find_one({"id": oid}, {"_id": 0})
    if not o:
        raise HTTPException(404, "sesi opname tidak ditemukan")
    return o


@router.patch("/opname/{oid}")
async def isi_opname(oid: str, body: OpnamePatchIn, _user=Depends(require_roles("petugas", "admin"))):
    """Simpan hasil hitung fisik (progresif). Hanya sesi draft."""
    db = get_db()
    o = await db.stok_opname.find_one({"id": oid}, {"_id": 0})
    if not o:
        raise HTTPException(404, "sesi opname tidak ditemukan")
    if o.get("status") != "draft":
        raise HTTPException(409, "sesi sudah selesai")
    fisik_map = {x.item_id: x.fisik for x in body.items}
    for it in o["items"]:
        if it["item_id"] in fisik_map:
            it["fisik"] = fisik_map[it["item_id"]]
    upd = {"items": o["items"]}
    if body.catatan is not None:
        upd["catatan"] = body.catatan
    await db.stok_opname.update_one({"id": oid}, {"$set": upd})
    return {"ok": True}


@router.post("/opname/{oid}/finalisasi")
async def finalisasi_opname(oid: str, user=Depends(require_roles("petugas", "admin"))):
    """Terapkan selisih sebagai transaksi penyesuaian. Saldo item disamakan ke fisik.
    Selisih dihitung ulang terhadap saldo SAAT INI (bukan snapshot), agar akurat."""
    db = get_db()
    o = await db.stok_opname.find_one({"id": oid}, {"_id": 0})
    if not o:
        raise HTTPException(404, "sesi opname tidak ditemukan")
    if o.get("status") != "draft":
        raise HTTPException(409, "sesi sudah selesai")
    now = datetime.now(timezone.utc)
    ids = [it["item_id"] for it in o["items"] if it.get("fisik") is not None]
    saldo = await _saldo_map(db, ids)
    n_sesuai = 0
    total_delta = 0.0
    for it in o["items"]:
        if it.get("fisik") is None:
            continue
        current = saldo.get(it["item_id"], 0)
        delta = float(it["fisik"]) - current
        it["sistem_akhir"] = _rapikan(current)
        it["selisih"] = _rapikan(delta)
        if delta != 0:
            await db.stok_transaksi.insert_one({
                "id": uuid.uuid4().hex,
                "item_id": it["item_id"], "item_nama": it["nama"], "satuan": it.get("satuan"),
                "jenis": "penyesuaian", "mutasi": delta,
                "tgl": now.date().isoformat(), "batch": None, "exp": None,
                "sumber": "opname", "catatan": f"opname {o['tgl']}", "ref_opname": oid,
                "petugas_id": user["id"], "created_at": now,
            })
            n_sesuai += 1
            total_delta += delta
    await db.stok_opname.update_one({"id": oid}, {"$set": {
        "status": "selesai", "items": o["items"], "selesai_at": now, "selesai_oleh": user["id"],
    }})
    return {"ok": True, "penyesuaian_dibuat": n_sesuai, "total_selisih": _rapikan(total_delta)}


@router.delete("/opname/{oid}")
async def hapus_opname(oid: str, _user=Depends(require_roles("admin"))):
    o = await get_db().stok_opname.find_one({"id": oid}, {"_id": 0, "status": 1})
    if not o:
        raise HTTPException(404, "tidak ditemukan")
    if o.get("status") == "selesai":
        raise HTTPException(409, "sesi selesai tak boleh dihapus (audit)")
    await get_db().stok_opname.delete_one({"id": oid})
    return {"ok": True}


# --------------------------------------------------------------- IMPORT EXCEL
class ImportIn(BaseModel):
    tipe: str                 # obat | alkes
    file_base64: str


def _norm(s):
    return "".join(str(s or "").lower().split())


def _exp_iso(v):
    from datetime import datetime as _dt
    if v is None or v == "":
        return None
    if hasattr(v, "date"):
        try:
            return v.date().isoformat()
        except Exception:
            pass
    s = str(v).strip()
    return s[:10] if s else None


@router.post("/import")
async def import_excel(body: ImportIn, user=Depends(require_roles("admin"))):
    """Impor Excel stok → item + stok awal (transaksi masuk). Idempoten: baris yang
    itemnya sudah ada (tipe+nama+sediaan) DILEWATI agar tak dobel."""
    if body.tipe not in TIPE:
        raise HTTPException(400, "tipe harus obat/alkes")
    import base64
    import io
    from openpyxl import load_workbook
    try:
        raw = base64.b64decode(body.file_base64)
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"gagal baca Excel: {e}")

    rows = list(ws.iter_rows(values_only=True))
    # cari baris header
    hdr_idx, header = None, None
    for i, r in enumerate(rows[:10]):
        cells = [_norm(c) for c in r]
        if ("namaproduk" in cells) or ("namaalat" in cells) or ("nama" in cells):
            hdr_idx, header = i, cells
            break
    if hdr_idx is None:
        raise HTTPException(400, "header tidak dikenali (butuh kolom Nama Produk/Nama Alat)")

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_nama = col("namaproduk", "namaalat", "nama")
    c_prod = col("produsen")
    c_kem = col("kemasan")
    c_sed = col("sediaan")
    c_kat = col("kategori")
    c_gauge = col("jarum", "gauge")
    c_qty = col("kuantiti")           # <-- stok saat ini (sesuai keputusan)
    c_exp = col("expired", "exp")
    c_ket = col("keterangan")

    db = get_db()
    now = datetime.now(timezone.utc)
    item_baru, item_reuse, tx_baru, tx_lewati = 0, 0, 0, 0
    contoh = []
    for r in rows[hdr_idx + 1:]:
        nama = r[c_nama] if c_nama is not None else None
        if not nama or not str(nama).strip():
            continue
        nama = str(nama).strip()
        sediaan = str(r[c_sed]).strip() if (c_sed is not None and r[c_sed] not in (None, "")) else None
        kemasan = str(r[c_kem]).strip() if (c_kem is not None and r[c_kem] not in (None, "")) else None
        satuan = kemasan or "unit"
        exp = _exp_iso(r[c_exp]) if c_exp is not None else None
        qty = 0.0
        if c_qty is not None:
            try:
                qty = float(r[c_qty] or 0)
            except (TypeError, ValueError):
                qty = 0.0

        # item: reuse bila (tipe+nama+sediaan) sudah ada, selain itu buat
        existing = await db.stok_item.find_one({"tipe": body.tipe, "nama": nama, "sediaan": sediaan}, {"_id": 0, "id": 1})
        if existing:
            item_id = existing["id"]
            item_reuse += 1
        else:
            item_id = uuid.uuid4().hex
            await db.stok_item.insert_one({
                "id": item_id, "tipe": body.tipe, "nama": nama, "satuan": satuan,
                "obat_id": None, "stok_minimum": None,
                "produsen": str(r[c_prod]).strip() if (c_prod is not None and r[c_prod] not in (None, "")) else None,
                "kemasan": kemasan, "sediaan": sediaan,
                "kategori": str(r[c_kat]).strip() if (c_kat is not None and r[c_kat] not in (None, "")) else None,
                "gauge": str(r[c_gauge]).strip() if (c_gauge is not None and r[c_gauge] not in (None, "")) else None,
                "keterangan": str(r[c_ket]).strip() if (c_ket is not None and r[c_ket] not in (None, "")) else None,
                "aktif": True, "created_by": user["id"], "created_at": now,
            })
            item_baru += 1

        # stok awal (transaksi masuk) — idempoten via import_key (aman diimpor ulang)
        if qty > 0:
            import_key = f"{body.tipe}|{nama}|{sediaan}|{exp}|{qty}"
            dup = await db.stok_transaksi.find_one({"import_key": import_key}, {"_id": 0, "id": 1})
            if dup:
                tx_lewati += 1
            else:
                await db.stok_transaksi.insert_one({
                    "id": uuid.uuid4().hex, "item_id": item_id, "item_nama": nama, "satuan": satuan,
                    "jenis": "masuk", "mutasi": qty, "tgl": now.date().isoformat(),
                    "batch": None, "exp": exp, "sumber": "stok awal", "catatan": "impor Excel",
                    "import_key": import_key, "petugas_id": user["id"], "created_at": now,
                })
                tx_baru += 1
                contoh.append({"nama": nama, "qty": _rapikan(qty), "exp": exp})
    return {"ok": True, "tipe": body.tipe, "item_baru": item_baru, "item_reuse": item_reuse,
            "stok_awal_dibuat": tx_baru, "stok_awal_dilewati": tx_lewati, "contoh": contoh[:5]}
