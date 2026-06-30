"""Rekap laporan bulanan — agregasi data tercatat menjadi ringkasan bulanan,
menggantikan rekap manual. v1: pelayanan (KESWAN) per penyakit/wilayah/petugas/
metode/prognosa + pemakaian obat + mutasi ternak (termasuk kematian) + peternak/
pendaftaran baru. Agregasi di Python (volume bulanan kecil). Mudah diperluas saat
kategori pelayanan lain ditambahkan."""
from datetime import datetime, timezone
from typing import Optional
from collections import Counter, defaultdict
import io

from fastapi import APIRouter, Depends
from fastapi.responses import Response

from db import get_db
from auth import require_roles

router = APIRouter(prefix="/laporan", tags=["laporan"])


def _range_str(tahun: int, bulan: int):
    return f"{tahun:04d}-{bulan:02d}-01", f"{tahun:04d}-{bulan:02d}-31"


def _range_dt(tahun: int, bulan: int):
    dari = datetime(tahun, bulan, 1, tzinfo=timezone.utc)
    sampai = datetime(tahun + 1, 1, 1, tzinfo=timezone.utc) if bulan == 12 else datetime(tahun, bulan + 1, 1, tzinfo=timezone.utc)
    return dari, sampai


@router.get("/bulanan")
async def laporan_bulanan(
    tahun: int,
    bulan: int,
    kalurahan_id: Optional[str] = None,
    _user=Depends(require_roles("petugas", "admin")),
):
    return await _hitung_rekap(get_db(), tahun, bulan, kalurahan_id)


def _num(x):
    """Tampilkan bilangan bulat tanpa desimal, selain itu apa adanya."""
    try:
        f = float(x)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return x


def _rincian_kategori(pels):
    """Ringkasan turunan per kategori dari field detail tiap pelayanan."""
    from collections import defaultdict

    def det(p):
        return p.get("detail") or {}

    def tabel(judul, kolom, baris):
        return {"judul": judul, "kolom": kolom, "baris": baris}

    def grup(rows, field, kolom_label):
        c = defaultdict(int)
        for d in rows:
            c[d.get(field) or "(tak diisi)"] += 1
        baris = [[k, v] for k, v in sorted(c.items(), key=lambda kv: -kv[1])]
        return tabel(f"Per {kolom_label}", [kolom_label.capitalize(), "Jumlah"], baris)

    out = {}
    by = defaultdict(list)
    for p in pels:
        by[det(p).get("kategori")].append(det(p))

    # VAKSINASI: total dosis + per jenis vaksin (kegiatan & dosis)
    if by.get("VAKSINASI"):
        rows = by["VAKSINASI"]
        per = defaultdict(lambda: {"kegiatan": 0, "dosis": 0.0})
        total = 0.0
        for d in rows:
            j = d.get("jenis_vaksin") or "(tak diisi)"
            try:
                dz = float(d.get("jumlah_dosis") or 0)
            except (TypeError, ValueError):
                dz = 0.0
            per[j]["kegiatan"] += 1
            per[j]["dosis"] += dz
            total += dz
        baris = [[j, v["kegiatan"], _num(v["dosis"])] for j, v in sorted(per.items(), key=lambda kv: -kv[1]["kegiatan"])]
        out["VAKSINASI"] = {
            "ringkas": {"Total kegiatan": len(rows), "Total dosis": _num(total)},
            "tabel": [tabel("Per jenis vaksin", ["Vaksin", "Kegiatan", "Dosis"], baris)],
        }

    # PKB: hasil bunting/tidak/meragukan
    if by.get("PKB"):
        rows = by["PKB"]
        hasil = defaultdict(int)
        for d in rows:
            hasil[d.get("hasil") or "(tak diisi)"] += 1
        out["PKB"] = {
            "ringkas": {"Total kegiatan": len(rows), "Bunting": hasil.get("Bunting", 0)},
            "tabel": [tabel("Per hasil", ["Hasil", "Jumlah"], [[k, v] for k, v in sorted(hasil.items(), key=lambda kv: -kv[1])])],
        }

    # IB: per pejantan/straw
    if by.get("IB"):
        rows = by["IB"]
        out["IB"] = {"ringkas": {"Total kegiatan": len(rows)}, "tabel": [grup(rows, "kode_pejantan", "pejantan")]}

    # GANGREP: per jenis gangguan
    if by.get("GANGREP"):
        rows = by["GANGREP"]
        out["GANGREP"] = {"ringkas": {"Total kegiatan": len(rows)}, "tabel": [grup(rows, "jenis_gangguan", "jenis gangguan")]}

    # LAB: per pemeriksaan
    if by.get("LAB"):
        rows = by["LAB"]
        out["LAB"] = {"ringkas": {"Total kegiatan": len(rows)}, "tabel": [grup(rows, "pemeriksaan", "pemeriksaan")]}

    # PEMBINAAN: total peserta
    if by.get("PEMBINAAN"):
        rows = by["PEMBINAAN"]
        peserta = 0.0
        for d in rows:
            try:
                peserta += float(d.get("jml_peserta") or 0)
            except (TypeError, ValueError):
                pass
        out["PEMBINAAN"] = {
            "ringkas": {"Total kegiatan": len(rows), "Total peserta": _num(peserta)},
            "tabel": [grup(rows, "topik", "topik")],
        }

    # KONSULTASI: per topik
    if by.get("KONSULTASI"):
        rows = by["KONSULTASI"]
        out["KONSULTASI"] = {"ringkas": {"Total kegiatan": len(rows)}, "tabel": [grup(rows, "topik", "topik")]}

    # ADUAN: per jenis aduan
    if by.get("ADUAN"):
        rows = by["ADUAN"]
        out["ADUAN"] = {"ringkas": {"Total kegiatan": len(rows)}, "tabel": [grup(rows, "jenis_aduan", "jenis aduan")]}

    # DESINFEKSI: per lokasi
    if by.get("DESINFEKSI"):
        rows = by["DESINFEKSI"]
        out["DESINFEKSI"] = {"ringkas": {"Total kegiatan": len(rows)}, "tabel": [grup(rows, "lokasi", "lokasi")]}

    return out


async def _hitung_rekap(db, tahun: int, bulan: int, kalurahan_id: Optional[str] = None):
    s_dari, s_sampai = _range_str(tahun, bulan)
    d_dari, d_sampai = _range_dt(tahun, bulan)

    # ---- pelayanan ----
    pflt = {"tgl": {"$gte": s_dari, "$lte": s_sampai}}
    if kalurahan_id:
        pflt["peternak.wilayah_id"] = kalurahan_id
    pels = await db.pelayanan.find(pflt, {"_id": 0}).to_list(5000)

    per_kategori, per_metode, per_prognosa, per_modalitas = Counter(), Counter(), Counter(), Counter()
    per_penyakit, per_wilayah, per_petugas = Counter(), Counter(), Counter()
    obat_agg = defaultdict(lambda: {"jumlah": 0.0, "satuan": ""})
    for p in pels:
        per_kategori[p.get("kategori") or "-"] += 1
        if p.get("metode_layanan"):
            per_metode[p["metode_layanan"]] += 1
        if p.get("modalitas"):
            per_modalitas[p["modalitas"]] += 1
        if p.get("prognosa"):
            per_prognosa[p["prognosa"]] += 1
        if p.get("penyakit_id"):
            per_penyakit[p["penyakit_id"]] += 1
        wid = (p.get("peternak") or {}).get("wilayah_id")
        if wid:
            per_wilayah[wid] += 1
        if p.get("petugas_id"):
            per_petugas[p["petugas_id"]] += 1
        for o in (p.get("obat") or []):
            nm = o.get("nama") or "?"
            obat_agg[nm]["jumlah"] += float(o.get("jumlah") or 0)
            obat_agg[nm]["satuan"] = o.get("satuan") or obat_agg[nm]["satuan"]

    penyakit_list = []
    for kode, n in per_penyakit.most_common():
        pen = await db.penyakit.find_one({"kode": kode}, {"_id": 0, "nama": 1})
        penyakit_list.append({"kode": kode, "nama": pen["nama"] if pen else kode, "jumlah": n})

    wilayah_list = []
    for wid, n in per_wilayah.most_common():
        w = await db.wilayah.find_one({"id": wid}, {"_id": 0, "nama": 1})
        wilayah_list.append({"kalurahan_id": wid, "nama": w["nama"] if w else wid, "jumlah": n})

    petugas_list = []
    for uid, n in per_petugas.most_common():
        u = await db.users.find_one({"id": uid}, {"_id": 0, "nama": 1})
        petugas_list.append({"petugas_id": uid, "nama": u["nama"] if u else uid, "jumlah": n})

    obat_list = sorted(
        [{"nama": k, "jumlah": round(v["jumlah"], 2), "satuan": v["satuan"]} for k, v in obat_agg.items()],
        key=lambda x: -x["jumlah"],
    )

    # ---- mutasi ternak (datetime) ----
    mut = await db.mutasi_ternak.find({"tgl": {"$gte": d_dari, "$lt": d_sampai}}, {"_id": 0}).to_list(5000)
    mutasi = Counter()
    for m in mut:
        mutasi[m.get("jenis") or "-"] += int(m.get("jumlah") or 1)

    pet_baru = await db.peternak.count_documents({"created_at": {"$gte": d_dari, "$lt": d_sampai}})
    daftar_baru = await db.pendaftaran.count_documents({"created_at": {"$gte": d_dari, "$lt": d_sampai}})
    daftar_konf = await db.pendaftaran.count_documents(
        {"created_at": {"$gte": d_dari, "$lt": d_sampai}, "status": "dikonfirmasi"}
    )

    return {
        "periode": {"tahun": tahun, "bulan": bulan, "dari": s_dari, "sampai": s_sampai},
        "filter_kalurahan_id": kalurahan_id,
        "pelayanan": {
            "total": len(pels),
            "per_kategori": dict(per_kategori),
            "rincian_kategori": _rincian_kategori(pels),
            "per_metode": dict(per_metode),
            "per_modalitas": dict(per_modalitas),
            "per_prognosa": dict(per_prognosa),
            "per_penyakit": penyakit_list,
            "per_wilayah": wilayah_list,
            "per_petugas": petugas_list,
        },
        "obat": obat_list,
        "ternak_mutasi": dict(mutasi),
        "peternak_baru": pet_baru,
        "pendaftaran": {"baru": daftar_baru, "dikonfirmasi": daftar_konf},
    }


# ---------------------------------------------------------------------------
# Ekspor Excel (.xlsx) multi-sheet — rapi, dekat dengan template laporan.
# ---------------------------------------------------------------------------
def _build_xlsx(rekap: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="0F6E56")
    TITLE = Font(bold=True, size=14)
    SUB = Font(color="666666")

    pel = rekap["pelayanan"]
    bulan_nama = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
                  "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    periode_txt = f"{bulan_nama[rekap['periode']['bulan']]} {rekap['periode']['tahun']}"

    def header_row(ws, row, cols):
        for j, c in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=j, value=c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="left")

    def autowidth(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 50)

    wb = Workbook()

    # --- Ringkasan ---
    ws = wb.active
    ws.title = "Ringkasan"
    ws["A1"] = "REKAP BULANAN PUSKESWAN"
    ws["A1"].font = TITLE
    ws["A2"] = f"Periode: {periode_txt}"
    ws["A2"].font = SUB
    r = 4
    header_row(ws, r, ["Ringkasan", "Nilai"]); r += 1
    for label, val in [
        ("Total pelayanan", pel["total"]),
        ("Peternak baru", rekap["peternak_baru"]),
        ("Pendaftaran baru", rekap["pendaftaran"]["baru"]),
        ("Pendaftaran dikonfirmasi", rekap["pendaftaran"]["dikonfirmasi"]),
        ("Kematian ternak", rekap["ternak_mutasi"].get("mati", 0)),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1
    for judul, data in [
        ("Per kategori", pel["per_kategori"]),
        ("Per modalitas", pel.get("per_modalitas", {})),
        ("Per metode", pel["per_metode"]),
        ("Per prognosa", pel["per_prognosa"]),
        ("Mutasi ternak", rekap["ternak_mutasi"]),
    ]:
        header_row(ws, r, [judul, "Jumlah"]); r += 1
        for k, v in (data or {}).items():
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            r += 1
        r += 1
    autowidth(ws)

    # --- sheet tabel ---
    def sheet_tabel(nama, kolom, baris):
        s = wb.create_sheet(nama)
        header_row(s, 1, kolom)
        for i, b in enumerate(baris, start=2):
            for j, c in enumerate(b, start=1):
                s.cell(row=i, column=j, value=c)
        autowidth(s)

    sheet_tabel("Penyakit", ["Kode", "Nama", "Jumlah"],
                [[x["kode"], x["nama"], x["jumlah"]] for x in pel["per_penyakit"]])
    sheet_tabel("Wilayah", ["Kalurahan", "Jumlah"],
                [[x["nama"], x["jumlah"]] for x in pel["per_wilayah"]])
    sheet_tabel("Petugas", ["Petugas", "Jumlah"],
                [[x["nama"], x["jumlah"]] for x in pel["per_petugas"]])
    sheet_tabel("Obat", ["Obat", "Jumlah", "Satuan"],
                [[x["nama"], x["jumlah"], x["satuan"]] for x in rekap["obat"]])

    # --- Rincian Kategori (turunan dari detail) ---
    rinc = pel.get("rincian_kategori") or {}
    if rinc:
        s = wb.create_sheet("Rincian Kategori")
        r = 1
        bulan_label = {
            "VAKSINASI": "Vaksinasi", "PKB": "PKB (kebuntingan)", "IB": "Inseminasi buatan",
            "GANGREP": "Gangguan reproduksi", "LAB": "Laboratorium", "PEMBINAAN": "Pembinaan",
            "KONSULTASI": "Konsultasi", "ADUAN": "Aduan", "DESINFEKSI": "Desinfeksi",
        }
        for kat, isi in rinc.items():
            c = s.cell(row=r, column=1, value=bulan_label.get(kat, kat))
            c.font = TITLE
            r += 1
            ringkas = isi.get("ringkas") or {}
            if ringkas:
                header_row(s, r, ["Ringkasan", "Nilai"]); r += 1
                for k, v in ringkas.items():
                    s.cell(row=r, column=1, value=k)
                    s.cell(row=r, column=2, value=v)
                    r += 1
                r += 1
            for tb in isi.get("tabel") or []:
                header_row(s, r, tb["kolom"]); r += 1
                for baris in tb["baris"]:
                    for j, val in enumerate(baris, start=1):
                        s.cell(row=r, column=j, value=val)
                    r += 1
                r += 1
            r += 1
        autowidth(s)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/bulanan/xlsx")
async def laporan_bulanan_xlsx(
    tahun: int,
    bulan: int,
    kalurahan_id: Optional[str] = None,
    _user=Depends(require_roles("petugas", "admin")),
):
    rekap = await _hitung_rekap(get_db(), tahun, bulan, kalurahan_id)
    data = _build_xlsx(rekap)
    fname = f"rekap-{tahun:04d}-{bulan:02d}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
